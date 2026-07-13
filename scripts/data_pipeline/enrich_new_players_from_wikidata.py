#!/usr/bin/env python3
"""
Enrich the late-replacement players with Wikidata identity data, and apply the
pending goalkeeper stat columns to the player_stats sheet.

For every active player missing place_of_birth/birth_country:
1. Search Wikidata by display name.
2. Accept a candidate only if its date of birth (P569) exactly matches ours
   (from the Wikipedia squads import) — the same conservative rule the
   original Step 3 pipeline used.
3. Pull place of birth (P19), its country (P17), and the QID; write them into
   the players sheet with data_confidence=wikidata_dob_match.
4. Anything unmatched lands in a review CSV instead of being guessed.

Also inserts gk_saves / gk_actions_inside_box / gk_actions_outside_box columns
into player_stats from data/processed/workbook_import/player_stats_goalkeeping_columns.csv
(the import file produced with the FIFA Goalkeeping-tab collection).

Backs up the workbook first; writes replication CSVs for the Google Sheets master.
"""

from __future__ import annotations

import csv
import shutil
import time
from datetime import date
from pathlib import Path

import requests
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
WB_PATH = REPO_ROOT / "data/master/world_cup_2026_player_map_master.xlsx"
OUT = REPO_ROOT / "data/processed/workbook_import/post_import_fixes"
TODAY = date.today().isoformat()

API = "https://www.wikidata.org/w/api.php"
HEADERS = {
    "User-Agent": "WorldCupPlayerTracker/0.1 (non-commercial research; torqueda@andrew.cmu.edu)"
}


def api_get(params: dict) -> dict:
    params = {"format": "json", **params}
    response = requests.get(API, params=params, headers=HEADERS, timeout=30)
    response.raise_for_status()
    return response.json()


def entity_claim_ids(entity: dict, prop: str) -> list[str]:
    out = []
    for claim in entity.get("claims", {}).get(prop, []):
        value = claim.get("mainsnak", {}).get("datavalue", {}).get("value")
        if isinstance(value, dict) and "id" in value:
            out.append(value["id"])
    return out


def entity_dob(entity: dict) -> str | None:
    for claim in entity.get("claims", {}).get("P569", []):
        value = claim.get("mainsnak", {}).get("datavalue", {}).get("value", {})
        timestr = value.get("time")  # e.g. +1997-05-10T00:00:00Z
        if timestr and value.get("precision", 0) >= 11:
            return timestr[1:11]
    return None


LABEL_CACHE: dict[str, dict] = {}


def get_entities(ids: list[str]) -> dict[str, dict]:
    missing = [qid for qid in ids if qid not in LABEL_CACHE]
    for i in range(0, len(missing), 50):
        chunk = missing[i : i + 50]
        data = api_get({"action": "wbgetentities", "ids": "|".join(chunk), "props": "claims|labels"})
        LABEL_CACHE.update(data.get("entities", {}))
        time.sleep(0.3)
    return {qid: LABEL_CACHE[qid] for qid in ids if qid in LABEL_CACHE}


def label_of(qid: str) -> str | None:
    entity = LABEL_CACHE.get(qid) or get_entities([qid]).get(qid)
    if not entity:
        return None
    return entity.get("labels", {}).get("en", {}).get("value")


def find_player(name: str, dob: str) -> dict | None:
    search = api_get({"action": "wbsearchentities", "search": name, "language": "en", "type": "item", "limit": 8})
    ids = [hit["id"] for hit in search.get("search", [])]
    if not ids:
        return None
    entities = get_entities(ids)
    for qid in ids:
        entity = entities.get(qid)
        if not entity:
            continue
        if entity_dob(entity) == dob:
            pob_ids = entity_claim_ids(entity, "P19")
            pob_label, country_label = None, None
            if pob_ids:
                pob_entity = get_entities([pob_ids[0]]).get(pob_ids[0])
                pob_label = label_of(pob_ids[0])
                if pob_entity:
                    country_ids = entity_claim_ids(pob_entity, "P17")
                    if country_ids:
                        country_label = label_of(country_ids[0])
            return {"qid": qid, "place_of_birth": pob_label, "birth_country": country_label}
    return None


def main() -> None:
    backup = WB_PATH.with_name(f"world_cup_2026_player_map_master.backup_{TODAY}_enrich.xlsx")
    shutil.copy2(WB_PATH, backup)
    print(f"Backup: {backup}")

    wb = load_workbook(WB_PATH)
    OUT.mkdir(parents=True, exist_ok=True)

    # --- 1. Wikidata enrichment -------------------------------------------
    p_ws = wb["players"]
    headers = [c.value for c in next(p_ws.iter_rows(min_row=1, max_row=1))]
    col = {name: headers.index(name) for name in headers if name}

    enriched, misses = [], []
    for row in p_ws.iter_rows(min_row=2):
        birth_country = row[col["birth_country"]].value
        place = row[col["place_of_birth"]].value
        dob_cell = row[col["date_of_birth"]].value
        if birth_country and place:
            continue
        name = str(row[col["display_name"]].value or "")
        dob = str(dob_cell)[:10] if dob_cell else None
        if not name or not dob:
            continue
        print(f"Searching Wikidata: {name} ({dob}) ...", end=" ")
        try:
            match = find_player(name, dob)
        except requests.RequestException as exc:
            print(f"ERROR {exc}")
            misses.append({"player_id": row[col["player_id"]].value, "display_name": name, "dob": dob, "reason": f"api error: {exc}"})
            continue
        time.sleep(0.4)
        if not match:
            print("no DOB-confirmed match")
            misses.append({"player_id": row[col["player_id"]].value, "display_name": name, "dob": dob, "reason": "no candidate with matching date of birth"})
            continue
        print(f"{match['qid']} | {match['place_of_birth']}, {match['birth_country']}")
        if match["place_of_birth"]:
            row[col["place_of_birth"]].value = match["place_of_birth"]
        if match["birth_country"]:
            row[col["birth_country"]].value = match["birth_country"]
        if not row[col["wikidata_id"]].value:
            row[col["wikidata_id"]].value = match["qid"]
        row[col["bio_source_url"]].value = f"https://www.wikidata.org/wiki/{match['qid']}"
        row[col["data_confidence"]].value = "wikidata_dob_match"
        note = f"birthplace enriched from Wikidata on {TODAY} (DOB-confirmed match)"
        existing = row[col["notes"]].value
        row[col["notes"]].value = f"{existing}; {note}" if existing else note
        enriched.append(
            {
                "player_id": row[col["player_id"]].value,
                "display_name": name,
                "wikidata_id": match["qid"],
                "place_of_birth": match["place_of_birth"],
                "birth_country": match["birth_country"],
                "bio_source_url": f"https://www.wikidata.org/wiki/{match['qid']}",
            }
        )

    # --- 2. Goalkeeper columns into player_stats ---------------------------
    gk_rows = list(csv.DictReader(open(REPO_ROOT / "data/processed/workbook_import/player_stats_goalkeeping_columns.csv", encoding="utf-8-sig")))
    gk_by_pid = {r["player_id"]: r for r in gk_rows if r["player_id"]}
    ps_ws = wb["player_stats"]
    ps_headers = [c.value for c in next(ps_ws.iter_rows(min_row=1, max_row=1))]
    added_cols = 0
    for name in ("gk_saves", "gk_actions_inside_box", "gk_actions_outside_box"):
        if name not in ps_headers:
            ps_ws.cell(row=1, column=len(ps_headers) + 1 + added_cols, value=name)
            added_cols += 1
    ps_headers = [c.value for c in next(ps_ws.iter_rows(min_row=1, max_row=1))]
    ps_col = {name: ps_headers.index(name) for name in ps_headers if name}
    filled = 0
    for row in ps_ws.iter_rows(min_row=2):
        pid = str(row[ps_col["player_id"]].value or "")
        gk = gk_by_pid.get(pid)
        if gk:
            row[ps_col["gk_saves"]].value = int(gk["gk_saves"])
            row[ps_col["gk_actions_inside_box"]].value = int(gk["gk_actions_inside_box"])
            row[ps_col["gk_actions_outside_box"]].value = int(gk["gk_actions_outside_box"])
            filled += 1
    print(f"\nGK stat rows filled: {filled} (columns added: {added_cols})")

    # --- 3. change_log ------------------------------------------------------
    wb["change_log"].append(
        [
            f"chg_{TODAY}_enrichment",
            TODAY,
            "ALL",
            "",
            "bulk_enrichment",
            "place_of_birth, birth_country, wikidata_id, gk stats",
            "",
            "",
            "https://www.wikidata.org/",
            "claude",
            f"Wikidata DOB-confirmed birthplace enrichment for {len(enriched)} late replacements ({len(misses)} to manual review); goalkeeper stat columns imported for {filled} keepers.",
        ]
    )

    wb.save(WB_PATH)
    print(f"Saved {WB_PATH}")

    def write(name: str, rows: list[dict], fieldnames: list[str]) -> None:
        path = OUT / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f"{path}: {len(rows)} rows")

    write("player_birth_enrichment.csv", enriched, ["player_id", "display_name", "wikidata_id", "place_of_birth", "birth_country", "bio_source_url"])
    write("player_birth_enrichment_misses.csv", misses, ["player_id", "display_name", "dob", "reason"])


if __name__ == "__main__":
    main()
