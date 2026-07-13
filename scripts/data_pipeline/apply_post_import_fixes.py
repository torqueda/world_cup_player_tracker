#!/usr/bin/env python3
"""
Post-import fixes for the 2026 final-dataset import (2026-07-12/13).

Applies directly to the local master xlsx (after writing a timestamped
backup next to it) and writes replication CSVs to
data/processed/workbook_import/post_import_fixes/ so the same edits can be
made in the Google Sheets master:

1. Remap the 12 placeholder club_ids in player_club_at_callup to the
   canonical clubs the user identified (Bournemouth -> AFC Bournemouth, ...).
2. Fill caps_pre_tournament / goals_pre_tournament for the 27 new squad
   entries from the Wikipedia squads scrape.
3. Note the FIFA alias on Mohammad Taha's player row (FIFA lists him as
   "Mohammad Abughoush").
4. Append the five collection-source rows to `sources`, shaped to the
   sheet's actual schema.
5. Append a change_log row documenting the import and these fixes.
"""

from __future__ import annotations

import csv
import shutil
import unicodedata
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
WB_PATH = REPO_ROOT / "data/master/world_cup_2026_player_map_master.xlsx"
RECON = REPO_ROOT / "data/processed/final_roster_reconciliation"
OUT = REPO_ROOT / "data/processed/workbook_import/post_import_fixes"
TODAY = date.today().isoformat()

# Placeholder club name (from the import) -> canonical club_name in `clubs`.
CLUB_ALIAS_TO_CANONICAL = {
    "Bournemouth": "AFC Bournemouth",
    "Lyon": "Olympique Lyonnais",
    "Chelsea": "Chelsea F.C.",
    "Mainz 05": "1. FSV Mainz 05",
    "Colorado Springs Switchbacks FC": "Colorado Springs Switchbacks",
    "Kalba": "Kalba FC",
    "Al-Karma": "Al-Karma SC",
    "Al-Hussein": "Al-Hussein SC (Irbid)",
    "Sunderland": "Sunderland A.F.C.",
    "Al-Hilal": "Al Hilal SFC",
    "Manchester United": "Manchester United F.C.",
    "Navbahor Namangan": "PFC Navbahor Namangan",
}

NEW_SOURCE_ROWS = [
    # source_id, source_type, publisher, url, title, accessed_at, published_at, reliability_tier, notes
    ("wiki_squads_2026", "roster", "Wikipedia", "https://en.wikipedia.org/wiki/2026_FIFA_World_Cup_squads",
     "2026 FIFA World Cup squads", TODAY, "", "secondary_verified",
     "Final 26-man rosters, shirt numbers, coaches, replacements. Snapshot: data/raw/wikipedia_2026_squads_snapshot.html"),
    ("wiki_officials_2026", "officials", "Wikipedia", "https://en.wikipedia.org/wiki/2026_FIFA_World_Cup_officials",
     "2026 FIFA World Cup officials", TODAY, "", "secondary_verified",
     "Referees, assistants, VARs. Snapshot: data/raw/wikipedia_2026_officials_snapshot.html"),
    ("fifa_player_stats_2026", "stats", "FIFA", "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/statistics/player-statistics",
     "FIFA.com player statistics (through quarterfinals)", TODAY, "", "primary_official",
     "Goals/assists/minutes for all registered players; cards from Discipline tab."),
    ("fifa_team_stats_2026", "stats", "FIFA", "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/statistics/team-statistics",
     "FIFA.com team statistics (through quarterfinals)", TODAY, "", "primary_official",
     "Team goals, assists, xG, possession."),
    ("fifa_fixtures_2026", "matches", "FIFA", "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/scores-fixtures",
     "FIFA.com scores & fixtures", TODAY, "", "primary_official",
     "All match results through the quarterfinals plus remaining schedule. Snapshot: data/raw/fifa_scores_fixtures_snapshot.txt"),
]


def fold(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.encode("ascii", "ignore").decode("ascii").casefold()
    return re.sub(r"\s+", " ", text).strip()


def sheet_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=False))
    headers = [cell.value for cell in rows[0]]
    return headers, rows[1:]


def col_index(headers, name):
    return headers.index(name)


def main() -> None:
    backup = WB_PATH.with_name(f"world_cup_2026_player_map_master.backup_{TODAY}.xlsx")
    shutil.copy2(WB_PATH, backup)
    print(f"Backup written: {backup}")

    wb = load_workbook(WB_PATH)
    OUT.mkdir(parents=True, exist_ok=True)
    replication: dict[str, list[list[object]]] = {"callup_club_id_fixes": [], "squad_caps_fills": []}

    # --- canonical club lookup -------------------------------------------
    clubs_ws = wb["clubs"]
    c_headers, c_rows = sheet_as_dicts(clubs_ws)
    name_i = col_index(c_headers, "club_name")
    id_i = col_index(c_headers, "club_id")
    club_id_by_name = {fold(str(r[name_i].value)): str(r[id_i].value) for r in c_rows if r[name_i].value}

    canonical_ids = {}
    for alias, canonical in CLUB_ALIAS_TO_CANONICAL.items():
        cid = club_id_by_name.get(fold(canonical))
        if not cid:
            raise SystemExit(f"Canonical club not found in clubs sheet: {canonical!r}")
        canonical_ids[alias] = cid
        print(f"  {alias:34} -> {canonical:30} {cid}")

    # --- 1. fix callup club_ids ------------------------------------------
    new_callups = list(csv.DictReader(open(REPO_ROOT / "data/processed/workbook_import/player_club_at_callup_new_rows.csv", encoding="utf-8-sig")))
    placeholder_to_canonical = {}
    for row in new_callups:
        alias = row["club_name_at_source"]
        if alias in CLUB_ALIAS_TO_CANONICAL:
            placeholder_to_canonical[row["club_id"]] = (canonical_ids[alias], CLUB_ALIAS_TO_CANONICAL[alias])

    pc_ws = wb["player_club_at_callup"]
    pc_headers, pc_rows = sheet_as_dicts(pc_ws)
    pc_club_i = col_index(pc_headers, "club_id")
    pc_pid_i = col_index(pc_headers, "player_id")
    pc_notes_i = col_index(pc_headers, "notes")
    fixed = 0
    for r in pc_rows:
        cid = str(r[pc_club_i].value or "")
        if cid in placeholder_to_canonical:
            new_id, canonical_name = placeholder_to_canonical[cid]
            r[pc_club_i].value = new_id
            note = f"club_id remapped from placeholder to canonical '{canonical_name}' on {TODAY}"
            r[pc_notes_i].value = f"{r[pc_notes_i].value}; {note}" if r[pc_notes_i].value else note
            replication["callup_club_id_fixes"].append([str(r[pc_pid_i].value), cid, new_id, canonical_name])
            fixed += 1
    print(f"Callup club_id fixes applied: {fixed}")

    # --- 2. caps/goals for the 27 new squad entries -----------------------
    wiki = list(csv.DictReader(open(RECON / "wikipedia_final_squads.csv", encoding="utf-8-sig")))
    new_players = list(csv.DictReader(open(REPO_ROOT / "data/processed/workbook_import/players_new_rows.csv", encoding="utf-8-sig")))
    caps_by_name = {fold(r["player_name"]): (r.get("caps") or "", r.get("goals") or "") for r in wiki}
    caps_by_pid = {}
    for np_row in new_players:
        key = fold(np_row["display_name"])
        if key in caps_by_name:
            caps_by_pid[np_row["player_id"]] = caps_by_name[key]

    se_ws = wb["squad_entries"]
    se_headers, se_rows = sheet_as_dicts(se_ws)
    se_pid_i = col_index(se_headers, "player_id")
    se_caps_i = col_index(se_headers, "caps_pre_tournament")
    se_goals_i = col_index(se_headers, "goals_pre_tournament")
    filled = 0
    for r in se_rows:
        pid = str(r[se_pid_i].value or "")
        if pid in caps_by_pid and r[se_caps_i].value in (None, ""):
            caps, goals = caps_by_pid[pid]
            r[se_caps_i].value = int(caps) if str(caps).isdigit() else caps
            r[se_goals_i].value = int(goals) if str(goals).isdigit() else goals
            replication["squad_caps_fills"].append([pid, caps, goals])
            filled += 1
    print(f"Squad caps/goals filled: {filled}")

    # --- 3. FIFA alias note on Mohammad Taha ------------------------------
    p_ws = wb["players"]
    p_headers, p_rows = sheet_as_dicts(p_ws)
    p_pid_i = col_index(p_headers, "player_id")
    p_notes_i = col_index(p_headers, "notes")
    for r in p_rows:
        if str(r[p_pid_i].value) == "p_jor_mohammad_taha_fd5d6fec":
            alias_note = "FIFA lists this player as 'Mohammad Abughoush' (same person; confirmed by 26-player roster elimination and manual lookup)."
            existing = str(r[p_notes_i].value or "")
            if "Abughoush" not in existing:
                r[p_notes_i].value = f"{existing}; {alias_note}" if existing else alias_note
                print("Added FIFA alias note to Mohammad Taha.")

    # --- 4. sources rows ---------------------------------------------------
    s_ws = wb["sources"]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(min_row=1, max_row=1))]
    existing_ids = {str(row[0]) for row in s_ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    added_sources = 0
    for row in NEW_SOURCE_ROWS:
        if row[0] not in existing_ids:
            s_ws.append(list(row))
            added_sources += 1
    print(f"Source rows appended: {added_sources} (schema: {s_headers})")

    # --- 5. change_log row --------------------------------------------------
    cl_ws = wb["change_log"]
    next_id = cl_ws.max_row  # header is row 1, so max_row == count + 1; simple sequential id
    cl_ws.append(
        [
            f"chg_{TODAY}_final_dataset_import",
            TODAY,
            "ALL",
            "",
            "bulk_import",
            "final rosters, stats, matches, coaches, referees",
            "",
            "",
            "https://en.wikipedia.org/wiki/2026_FIFA_World_Cup_squads",
            "user+claude",
            f"Imported final-dataset collection of {TODAY}; remapped 12 placeholder club_ids to canonical clubs; filled caps/goals for 27 replacements; recorded FIFA alias for Mohammad Taha.",
        ]
    )
    print("Change_log row appended.")

    wb.save(WB_PATH)
    print(f"Saved workbook: {WB_PATH}")

    for name, rows in replication.items():
        path = OUT / f"{name}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            if name == "callup_club_id_fixes":
                w.writerow(["player_id", "old_club_id", "new_club_id", "canonical_club_name"])
            else:
                w.writerow(["player_id", "caps_pre_tournament", "goals_pre_tournament"])
            w.writerows(rows)
        print(f"Replication CSV: {path} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
