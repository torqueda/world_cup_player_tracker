#!/usr/bin/env python3
"""
Utilities for reading simple row-based Excel workbook sheets.

These helpers prefer a normal `openpyxl` install, but can also fall back to the
bundled Codex runtime packages on the same machine.
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any


def _ensure_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore

        return load_workbook
    except ImportError:
        pass

    runtime_root = Path.home() / ".cache" / "codex-runtimes"
    if runtime_root.exists():
        candidates: list[Path] = []
        for root_candidate in sorted(runtime_root.rglob("dependencies/python"), reverse=True):
            if not root_candidate.is_dir():
                continue
            candidates.append(root_candidate)
            candidates.extend(sorted(root_candidate.rglob("site-packages"), reverse=True))

        for candidate in candidates:
            if not candidate.is_dir():
                continue
            sys.path.insert(0, str(candidate))
            try:
                from openpyxl import load_workbook  # type: ignore

                return load_workbook
            except ImportError:
                sys.path.pop(0)

    raise SystemExit(
        "openpyxl is required to read the master workbook. "
        "Install it in your Python environment or run this script where the Codex bundled runtime is available."
    )


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def read_sheet_rows(workbook_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    load_workbook = _ensure_openpyxl()
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    header_cells = list(header_row or [])
    last_header_index = max(
        (index for index, value in enumerate(header_cells) if value not in (None, "")),
        default=-1,
    )
    if last_header_index < 0:
        return []

    headers = [header_cells[index] for index in range(last_header_index + 1)]
    rows: list[dict[str, Any]] = []

    for raw_row in iterator:
        values = list(raw_row or [])[: len(headers)]
        if all(value in (None, "") for value in values):
            continue

        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header in (None, ""):
                continue
            value = values[index] if index < len(values) else None
            row[str(header)] = normalize_cell_value(value)
        rows.append(row)

    return rows


def list_sheet_names(workbook_path: str | Path) -> list[str]:
    load_workbook = _ensure_openpyxl()
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=False)
    return list(workbook.sheetnames)
